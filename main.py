from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium_stealth import stealth
from selenium import webdriver
from openpyxl import load_workbook
import requests
from openpyxl.styles import Font, Alignment, Border, Side

from selenium.webdriver.firefox.options import Options

import time
from datetime import datetime, timedelta
from get_link_on_case import scrap_inf
from common import (
    PLAINTIFFS,
    DEFENDANTS,
    THIRDS,
    OTHERS,
    INN,
    DATE,
    NUMBERS_CASE,
    ESSENCE_OF_CASE,
    COURTS,
    IS_PROXIES,
)
from load_data_to_exel import load_data_to_excel

fn = "sber.xlsx"
wb = load_workbook(fn)
ws = wb["data"]

# очистка листа
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.value = None

headers = [
    "Номер дела",
    "Дата",
    "Суд",
    "Истец/ИНН",
    "Ответчик/ИНН",
    "Третьи лица",
    "Иные лица",
    "Суть дела",
]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True, size=10)  # Делает текст жирным
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

wb.save(fn)
wb.close()

IS_PROXIES = True # Меняем на False, если нет бана по ip

options = webdriver.ChromeOptions()
user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
options.add_argument(f"--user-agent={user_agent}")
if IS_PROXIES:
    options.add_argument('--proxy-server=145.255.30.241:8088')
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option("useAutomationExtension", False)
options.add_argument("--disable-blink-features=AutomationControlled")
driver = webdriver.Chrome(options=options)


URL = "https://kad.arbitr.ru"

driver.maximize_window()
driver.get(URL)
time.sleep(10)

# Начальная дата
start_date = datetime.strptime("01.06.2024", "%d.%m.%Y")
end_date = datetime.strptime("30.06.2024", "%d.%m.%Y")

current_date = start_date

while current_date < end_date:
    if current_date.strftime("%d.%m.%Y") != "01.06.2024":
        current_date += timedelta(days=2)
    time.sleep(2)
    data_first = current_date.strftime("%d.%m.%Y")
    data_second = (current_date + timedelta(days=1)).strftime("%d.%m.%Y")
    # Ввод данных на страницу
    data_input = driver.find_elements(
        By.CSS_SELECTOR, 'input[class="anyway_position_top g-ph"]'
    )
    if data_input:
        data_first_input = data_input[0]
        data_second_input = data_input[1]
    else:
        data_input = driver.find_elements(
            By.CSS_SELECTOR, 'input[class="anyway_position_top"]'
        )
        data_first_input = data_input[0]
        data_second_input = data_input[1]

    data_first_input.click()
    time.sleep(3)
    data_first_input.clear()
    time.sleep(1)
    data_first_input.click()
    time.sleep(1)
    data_first_input.send_keys(data_first)
    time.sleep(1)

    data_second_input.click()
    time.sleep(1)
    data_second_input.clear()
    time.sleep(1)
    data_second_input.click()
    time.sleep(1)
    data_second_input.send_keys(data_second)
    time.sleep(3)
    data_second_input.send_keys(Keys.RETURN)
    time.sleep(2)

    while True:
        # try:
            time.sleep(5)
            # Получение cookies из Selenium
            cookies = driver.get_cookies()
            session = requests.Session()
            for cookie in cookies:
                session.cookies.set(cookie["name"], cookie["value"])

            session.headers.update({"User-Agent": UserAgent().random})

            page_source = driver.page_source
            soup = BeautifulSoup(page_source, "lxml")

            scrap_inf(soup, session)

            load_data_to_excel(
                PLAINTIFFS,
                DEFENDANTS,
                THIRDS,
                OTHERS,
                INN,
                DATE,
                NUMBERS_CASE,
                ESSENCE_OF_CASE,
                COURTS,
                r"C:\Users\User\Desktop\Projects\pythonProject1\sber.xlsx",
            )

            time.sleep(6)
            print(INN)
            next_button = driver.find_element(By.CSS_SELECTOR, 'li[class="rarr"]')
            next_button.click()
            time.sleep(6)
        # except:
        #     print("Следующая страница недоступна, переходим к следующей.")
        #     break


driver.quit()
