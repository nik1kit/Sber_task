from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side


def load_data_to_excel(
    plaintiffs,
    defendants,
    thirds,
    others,
    inns,
    dates,
    numbers_case,
    essence_of_case,
    courts,
    filename,
):
    wb = load_workbook(filename)
    ws = wb["data"]

    # Находим следующую строку для записи данных
    next_row = ws.max_row + 1

    for row_index in range(len(numbers_case)):
        # Номер дела
        ws.cell(
            row=next_row + row_index, column=1, value=numbers_case[row_index]
        ).font = Font(size=10)
        ws.cell(row=next_row + row_index, column=1).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=next_row + row_index, column=1).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Дата
        ws.cell(row=next_row + row_index, column=2, value=dates[row_index]).font = Font(
            size=10
        )
        ws.cell(row=next_row + row_index, column=2).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=next_row + row_index, column=2).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Суд
        ws.cell(
            row=next_row + row_index, column=3, value=courts[row_index]
        ).font = Font(size=10)
        ws.cell(row=next_row + row_index, column=3).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=next_row + row_index, column=3).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Истец / ИНН
        ws.cell(
            row=next_row + row_index,
            column=4,
            value=f"{plaintiffs[row_index]} / {inns[row_index][0]}",
        ).font = Font(size=10)
        ws.cell(row=next_row + row_index, column=4).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=next_row + row_index, column=4).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Ответчик / ИНН
        ws.cell(
            row=next_row + row_index,
            column=5,
            value=f"{defendants[row_index]} / {inns[row_index][1]}",
        ).font = Font(size=10)
        ws.cell(row=next_row + row_index, column=5).alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=next_row + row_index, column=5).border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Третьи лица
        if thirds[row_index]:
            ws.cell(
                row=next_row + row_index, column=6, value=", ".join(thirds[row_index])
            ).font = Font(size=10)
            ws.cell(row=next_row + row_index, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=6).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            ws.cell(row=next_row + row_index, column=6, value="-").font = Font(size=10)
            ws.cell(row=next_row + row_index, column=6).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=6).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Иные лица
        if others[row_index]:
            ws.cell(
                row=next_row + row_index, column=7, value=", ".join(others[row_index])
            ).font = Font(size=10)
            ws.cell(row=next_row + row_index, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=7).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            ws.cell(row=next_row + row_index, column=7, value="-").font = Font(size=10)
            ws.cell(row=next_row + row_index, column=7).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=7).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

        # Суть дела
        if essence_of_case[row_index]:
            ws.cell(
                row=next_row + row_index,
                column=8,
                value=", ".join(essence_of_case[row_index]),
            ).font = Font(size=10)
            ws.cell(row=next_row + row_index, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=8).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
        else:
            ws.cell(row=next_row + row_index, column=8, value="-").font = Font(size=10)
            ws.cell(row=next_row + row_index, column=8).alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws.cell(row=next_row + row_index, column=8).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    plaintiffs.clear()
    defendants.clear()
    thirds.clear()
    others.clear()
    inns.clear()
    dates.clear()
    numbers_case.clear()
    essence_of_case.clear()
    courts.clear()

    wb.save(filename)
    wb.close()
